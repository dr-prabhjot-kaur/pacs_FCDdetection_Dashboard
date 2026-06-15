#!/usr/bin/env python3
"""Convert the dashboard CSV into a formatted .xlsx (presentation copy)."""
import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
STATUS_COLS = {"meld_classifier", "meld_graph", "nnunet", "job_status"}
TEXT_COLS = {"job_id", "mrn"}
FILLS = {
    "DONE":     PatternFill("solid", fgColor="C6EFCE"),
    "COMPLETE": PatternFill("solid", fgColor="C6EFCE"),
    "FAILED":   PatternFill("solid", fgColor="FFC7CE"),
    "RUNNING":  PatternFill("solid", fgColor="FFEB9C"),
    "PENDING":  PatternFill("solid", fgColor="E7E6E6"),
}
FONTS = {
    "DONE":     Font(name=FONT, color="006100"),
    "COMPLETE": Font(name=FONT, color="006100"),
    "FAILED":   Font(name=FONT, color="9C0006"),
    "RUNNING":  Font(name=FONT, color="9C6500"),
    "PENDING":  Font(name=FONT, color="3F3F3F"),
}


def convert(in_csv, out_xlsx):
    with open(in_csv, newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise SystemExit(f"empty CSV: {in_csv}")
    header, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "FCD dashboard"

    hdr_font = Font(name=FONT, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    for c, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, line in enumerate(data, start=2):
        for c, name in enumerate(header, 1):
            val = line[c - 1] if c - 1 < len(line) else ""
            cell = ws.cell(row=r, column=c)
            if name in TEXT_COLS:
                cell.value = val
                cell.number_format = "@"
            else:
                cell.value = val
            cell.font = Font(name=FONT)
            if name in STATUS_COLS and val in FILLS:
                cell.fill = FILLS[val]
                cell.font = FONTS[val]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(data) + 1}"
    for c, name in enumerate(header, 1):
        width = max([len(name)] + [len(line[c - 1]) for line in data
                                   if c - 1 < len(line)] + [6])
        ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 40)

    tmp = out_xlsx + ".tmp"
    wb.save(tmp)
    os.replace(tmp, out_xlsx)
    print(f"wrote {out_xlsx}  ({len(data)} rows)")


if __name__ == "__main__":
    in_csv = sys.argv[1] if len(sys.argv) > 1 else "fcd_dashboard.csv"
    out_xlsx = (sys.argv[2] if len(sys.argv) > 2
                else os.path.splitext(in_csv)[0] + ".xlsx")
    convert(in_csv, out_xlsx)
